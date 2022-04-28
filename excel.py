import copy

import openpyxl
from openpyxl.utils import get_column_letter


def get_sheet_max_row(wb, sheet_name):
    sheet = wb[sheet_name]
    return sheet.max_row


def write_xlsx(path: str, excel_sheet_datas: list):
    workbook = openpyxl.Workbook()

    for i in range(0, len(excel_sheet_datas), 1):
        s = excel_sheet_datas[i]
        sheet_title = s['title']
        info = s['info']
        data = s['data']
        sheet = workbook.create_sheet(sheet_title, index=i)
        sheet.column_dimensions['A'].width = 30
        # 添加表头（不需要表头可以不用加）
        data.insert(0, list(info))

        for row_index, row_item in enumerate(data):
            for col_index, col_item in enumerate(row_item):
                sheet.cell(row=row_index + 1, column=col_index + 1, value=col_item)

    workbook.save(path)
    print('写入成功: ' + path)


def get_xlsx_row_detail(path: str) -> list:
    wb = openpyxl.load_workbook(path)
    sheet_names = wb.sheetnames
    sheet_names.sort()
    result = []
    for sheet_name in sheet_names:
        row = get_sheet_max_row(wb, sheet_name)
        row_num = row - 1
        result.append([sheet_name, row_num])
    return result


def get_xlsx_total_rows(path: str):
    wb = openpyxl.load_workbook(path)
    sheet_names = wb.sheetnames
    sheet_names.sort()
    total_row = 0
    for sheet_name in sheet_names:
        row = get_sheet_max_row(wb, sheet_name)
        total_row += row - 1
    return total_row


def copy_xlsx(path: str, save_path: str, copy_image=True):
    wb = openpyxl.load_workbook(path)
    wb2 = openpyxl.Workbook()

    sheet_names = wb.sheetnames
    sheet_names.sort()
    for sheet_name in sheet_names:
        print(sheet_name)
        sheet = wb[sheet_name]
        sheet2 = wb2.create_sheet(sheet_name)

        # tab颜色
        sheet2.sheet_properties.tabColor = sheet.sheet_properties.tabColor

        # 开始处理合并单元格形式为“(<CellRange A1：A4>,)，替换掉(<CellRange 和 >,)' 找到合并单元格
        wm = list(sheet.merged_cells)
        if len(wm) > 0:
            for i in range(0, len(wm)):
                cell2 = str(wm[i]).replace('(<CellRange ', '').replace('>,)', '')
                sheet2.merge_cells(cell2)

        for i, row in enumerate(sheet.iter_rows()):
            sheet2.row_dimensions[i + 1].height = sheet.row_dimensions[i + 1].height
            for j, cell in enumerate(row):
                sheet2.column_dimensions[get_column_letter(j + 1)].width = sheet.column_dimensions[
                    get_column_letter(j + 1)].width
                sheet2.cell(row=i + 1, column=j + 1, value=cell.value)

                # 设置单元格格式
                source_cell = sheet.cell(i + 1, j + 1)
                target_cell = sheet2.cell(i + 1, j + 1)
                target_cell.fill = copy.copy(source_cell.fill)
                if source_cell.has_style:
                    target_cell._style = copy.copy(source_cell._style)
                    target_cell.font = copy.copy(source_cell.font)
                    target_cell.border = copy.copy(source_cell.border)
                    target_cell.fill = copy.copy(source_cell.fill)
                    target_cell.number_format = copy.copy(source_cell.number_format)
                    target_cell.protection = copy.copy(source_cell.protection)
                    target_cell.alignment = copy.copy(source_cell.alignment)

        if copy_image:
            for image in sheet._images:
                sheet2.add_image(image)

    if 'Sheet' in wb2.sheetnames:
        del wb2['Sheet']
    wb2.save(save_path)

    wb.close()
    wb2.close()

    print('Done.')