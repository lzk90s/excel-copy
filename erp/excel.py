import copy
from types import LambdaType

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


######################################################
# workbook
######################################################


def load_workbook(path: str) -> object:
    wb = openpyxl.load_workbook(path)
    wb.original_path = path
    return wb

def close_workbook(wb: Workbook):
    wb.close()
    pass


def save_workbook(wb, path):
    wb.save(path)


######################################################
# worksheet
######################################################

def get_all_worksheets(wb) -> list:
    return wb._sheets


def parse_worksheets(wb, sheet_parser: LambdaType) -> list:
    result = []
    for sheet_name in wb.sheetnames:
        r = sheet_parser(wb[sheet_name])
        # assert isinstance(r, list)
        result.append(r)
    return result


def sort_worksheets(wb, sort_key):
    wb._sheets.sort(key=sort_key)


def remove_worksheet_by_name(wb, sheet_name):
    if wb and sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])


def copy_workbook(path: str, save_path: str, copy_image=True):
    wb = load_workbook(path)
    wb2 = openpyxl.Workbook()

    for sheet_name in wb.sheetnames:
        print(sheet_name)
        sheet = wb[sheet_name]
        sheet2 = wb2.create_sheet(sheet_name)

        # tab颜色
        sheet2.sheet_properties.tabColor = sheet.sheet_properties.tabColor

        # 开始处理合并单元格形式为“(<CellRange A1：A4>,)，替换掉(<CellRange 和 >,)' 找到合并单元格
        wm = list(sheet.merged_cells)
        if len(wm) > 0:
            for i in range(0, len(wm)):
                cell2 = str(wm[i]).replace('(<CellRange ', '').replace('>,)', '')
                sheet2.merge_cells(cell2)

        for i, row in enumerate(sheet.iter_rows()):
            sheet2.row_dimensions[i + 1].height = sheet.row_dimensions[i + 1].height
            for j, cell in enumerate(row):
                sheet2.column_dimensions[get_column_letter(j + 1)].width = sheet.column_dimensions[
                    get_column_letter(j + 1)].width
                sheet2.cell(row=i + 1, column=j + 1, value=cell.value)

                # 设置单元格格式
                source_cell = sheet.cell(i + 1, j + 1)
                target_cell = sheet2.cell(i + 1, j + 1)
                target_cell.fill = copy.copy(source_cell.fill)
                if source_cell.has_style:
                    target_cell._style = copy.copy(source_cell._style)
                    target_cell.font = copy.copy(source_cell.font)
                    target_cell.border = copy.copy(source_cell.border)
                    target_cell.fill = copy.copy(source_cell.fill)
                    target_cell.number_format = copy.copy(source_cell.number_format)
                    target_cell.protection = copy.copy(source_cell.protection)
                    target_cell.alignment = copy.copy(source_cell.alignment)

        if copy_image:
            for image in sheet._images:
                sheet2.add_image(image)

    if 'Sheet' in wb2.sheetnames:
        del wb2['Sheet']
    wb2.save(save_path)

    wb.close()
    wb2.close()

    print('Done.')


def add_worksheet(wb, sheet_datas: list):
    for i in range(0, len(sheet_datas), 1):
        s = sheet_datas[i]
        sheet_name = s['sheet_name']
        head = s['head']
        data = s['data']
        column_dimensions = s['column_dimensions']

        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])

        ws = wb.create_sheet(sheet_name, index=i)
        if head:
            data.insert(0, list(head))

        for j in range(0, len(column_dimensions), 1):
            ws.column_dimensions[get_column_letter(j + 1)].width = column_dimensions[j]

        for row_index, row_item in enumerate(data):
            for col_index, col_item in enumerate(row_item):
                ws.cell(row=row_index + 1, column=col_index + 1, value=col_item)


######################################################
# cell
######################################################

def get_cell_value(ws, row: int, column: int):
    assert isinstance(ws, Worksheet)
    return ws.cell(row, column).value


def set_cell_value(ws: Worksheet, row: int, colum: int, value):
    ws.cell(row, colum).value = value


def get_row_value(ws, row):
    if row > ws.max_row:
        return None
    res = []
    for i in range(0, ws.max_cloumn):
        res.append(ws.cell(row, i))
    return res


######################################################
# mix
######################################################
def _get_column_idx_by_head(ws: Worksheet, name: str, head_row_idx=1):
    for i in range(1, ws.max_column + 1, 1):
        c = ws.cell(head_row_idx, i).value
        if c == name:
            return i
    return None


def count_column_non_blank_rows(ws: Worksheet, column_idx: int, ignore_head=True) -> int:
    count = 0
    begin = 2 if ignore_head else 1
    for i in range(begin, ws.max_row + 1, 1):
        v = ws.cell(i, column_idx).value
        if v is not None:
            count = count + 1
    return count
